import re
from pathlib import Path
from typing import Optional

import pandas as pd
from openpyxl import load_workbook


class ExcelFormat(object):
    map_engine = {'.xls': 'xlrd', '.xlsx': 'openpyxl'}

    def __init__(self, path: Path):
        self.df = pd.read_excel(path, engine=self.map_engine[path.suffix], header=None)
        self.write_file = Path(__file__).parent.joinpath(f'format_{path.name.split(".")[0]}.xlsx')
        self.format_percent_columns = [self.df.iloc[0, i] for i, v in enumerate(self.df.iloc[1]) if isinstance(v, int)]
        self.length = len(self.df)

    def excel_format(self) -> Optional:
        first_row = self.df.iloc[0]
        for column, value in first_row.items():
            if bool(re.search('PD|MJ', value)):
                insert_column = self.df.columns.get_loc(column)
                self.insert_columns(insert_column, column, value)
        self.df.to_excel(self.write_file, index=False, header=False, engine='openpyxl')
        self.format_percent()

    def insert_columns(self, insert_idx, column, value) -> Optional:
        empty_df = pd.DataFrame({f'empty_{column}_{i}': [''] * self.length for i in range(4)})
        valid_df = pd.DataFrame({f'有效_{column}': ['有效'] + ['Y'] * (self.length - 1)})
        label_df = pd.DataFrame({f'label_{column}': ['标签编码'] + [value] * (self.length - 1)})
        self.df = pd.concat(
            objs=[
                self.df.iloc[:, :insert_idx], label_df, empty_df, self.df.iloc[:, insert_idx], valid_df,
                self.df.iloc[:, insert_idx + 1:]
            ], axis=1, join='outer',
        )

    def format_percent(self) -> Optional:
        wb = load_workbook(self.write_file)
        ws = wb.active
        for col_idx, col in enumerate(ws.iter_cols(), start=1):
            if col[0].value in self.format_percent_columns:
                for cell in col[1:]:
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '0%'
        wb.save(self.write_file)

    @classmethod
    def format_all_file(cls) -> Optional:
        plt = Path(__file__).parent
        for file in plt.glob('*.xls*'):
            if file.name.find('format_') != -1:
                continue
            print(f'格式化: {file}')
            ExcelFormat(file).excel_format()


if __name__ == '__main__':
    ExcelFormat.format_all_file()
